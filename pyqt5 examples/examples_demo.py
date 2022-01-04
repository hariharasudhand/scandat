from openpyxl import load_workbook

workbook = load_workbook(filename="G:\python\env\SMWSED_sep02_2021.xlsx")
#print(workbook.sheetnames)
sheet = workbook.active
#print(sheet["B5"].value)
value = (sheet.cell(row=9, column=2).value)
print(value)
length = len(value)
print(length)
#for i in range ():


